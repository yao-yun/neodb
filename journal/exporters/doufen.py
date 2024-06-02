import os
from datetime import datetime

from django.conf import settings
from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook

from catalog.models import *
from common.utils import GenerateDateUUIDMediaFilePath
from journal.models import *


def _get_source_url(item):
    res = (
        item.external_resources.all()
        .filter(
            id_type__in=[
                IdType.DoubanBook,
                IdType.DoubanMovie,
                IdType.DoubanMusic,
                IdType.DoubanGame,
                IdType.DoubanDrama,
            ]
        )
        .first()
    )
    if not res:
        res = item.external_resources.all().first()
    return res.url if res else ""


def export_marks_task(user):
    user.preference.export_status["marks_pending"] = True
    user.preference.save(update_fields=["export_status"])
    filename = GenerateDateUUIDMediaFilePath(
        "f.xlsx", settings.MEDIA_ROOT + "/" + settings.EXPORT_FILE_PATH_ROOT
    )
    if not os.path.exists(os.path.dirname(filename)):
        os.makedirs(os.path.dirname(filename))
    heading = [
        "标题",
        "简介",
        "豆瓣评分",
        "链接",
        "创建时间",
        "我的评分",
        "标签",
        "评论",
        "NeoDB链接",
        "其它ID",
    ]
    wb = Workbook()
    # adding write_only=True will speed up but corrupt the xlsx and won't be importable
    for status, label in [
        (ShelfType.COMPLETE, "看过"),
        (ShelfType.PROGRESS, "在看"),
        (ShelfType.WISHLIST, "想看"),
    ]:
        ws = wb.create_sheet(title=label)
        shelf = user.shelf_manager.get_shelf(status)
        q = q_item_in_category(ItemCategory.Movie) | q_item_in_category(ItemCategory.TV)
        marks = shelf.members.all().filter(q).order_by("created_time")
        ws.append(heading)
        for mm in marks:
            mark = mm.mark
            movie = mark.item
            title = movie.title
            summary = (
                str(movie.year or "")
                + " / "
                + ",".join(movie.area or [])
                + " / "
                + ",".join(movie.genre or [])
                + " / "
                + ",".join(movie.director or [])
                + " / "
                + ",".join(movie.actor or [])
            )
            tags = ",".join(mark.tags)
            world_rating = (movie.rating / 2) if movie.rating else None
            timestamp = mark.created_time.strftime("%Y-%m-%d %H:%M:%S")
            my_rating = (mark.rating_grade / 2) if mark.rating_grade else None
            text = mark.comment_text
            source_url = _get_source_url(movie)
            url = movie.absolute_url
            line = [
                title,
                summary,
                world_rating,
                source_url,
                timestamp,
                my_rating,
                tags,
                text,
                url,
                movie.imdb,
            ]
            ws.append(line)

    for status, label in [
        (ShelfType.COMPLETE, "听过"),
        (ShelfType.PROGRESS, "在听"),
        (ShelfType.WISHLIST, "想听"),
    ]:
        ws = wb.create_sheet(title=label)
        shelf = user.shelf_manager.get_shelf(status)
        q = q_item_in_category(ItemCategory.Music)
        marks = shelf.members.all().filter(q).order_by("created_time")
        ws.append(heading)
        for mm in marks:
            mark = mm.mark
            album = mark.item
            title = album.title
            summary = (
                ",".join(album.artist)
                + " / "
                + (album.release_date.strftime("%Y") if album.release_date else "")
            )
            tags = ",".join(mark.tags)
            world_rating = (album.rating / 2) if album.rating else None
            timestamp = mark.created_time.strftime("%Y-%m-%d %H:%M:%S")
            my_rating = (mark.rating_grade / 2) if mark.rating_grade else None
            text = mark.comment_text
            source_url = _get_source_url(album)
            url = album.absolute_url
            line = [
                title,
                summary,
                world_rating,
                source_url,
                timestamp,
                my_rating,
                tags,
                text,
                url,
                album.barcode,
            ]
            ws.append(line)

    for status, label in [
        (ShelfType.COMPLETE, "读过"),
        (ShelfType.PROGRESS, "在读"),
        (ShelfType.WISHLIST, "想读"),
    ]:
        ws = wb.create_sheet(title=label)
        shelf = user.shelf_manager.get_shelf(status)
        q = q_item_in_category(ItemCategory.Book)
        marks = shelf.members.all().filter(q).order_by("created_time")
        ws.append(heading)
        for mm in marks:
            mark = mm.mark
            book = mark.item
            title = book.title
            summary = (
                ",".join(book.author or [])
                + " / "
                + str(book.pub_year or "")
                + " / "
                + (book.pub_house or "")
            )
            tags = ",".join(mark.tags)
            world_rating = (book.rating / 2) if book.rating else None
            timestamp = mark.created_time.strftime("%Y-%m-%d %H:%M:%S")
            my_rating = (mark.rating_grade / 2) if mark.rating_grade else None
            text = mark.comment_text
            source_url = _get_source_url(book)
            url = book.absolute_url
            line = [
                title,
                summary,
                world_rating,
                source_url,
                timestamp,
                my_rating,
                tags,
                text,
                url,
                book.isbn,
            ]
            ws.append(line)

    for status, label in [
        (ShelfType.COMPLETE, "玩过"),
        (ShelfType.PROGRESS, "在玩"),
        (ShelfType.WISHLIST, "想玩"),
    ]:
        ws = wb.create_sheet(title=label)
        shelf = user.shelf_manager.get_shelf(status)
        q = q_item_in_category(ItemCategory.Game)
        marks = shelf.members.all().filter(q).order_by("created_time")
        ws.append(heading)
        for mm in marks:
            mark = mm.mark
            game = mark.item
            title = game.title
            summary = (
                ",".join(game.genre or [])
                + " / "
                + ",".join(game.platform or [])
                + " / "
                + (game.release_date.strftime("%Y-%m-%d") if game.release_date else "")
            )
            tags = ",".join(mark.tags)
            world_rating = (game.rating / 2) if game.rating else None
            timestamp = mark.created_time.strftime("%Y-%m-%d %H:%M:%S")
            my_rating = (mark.rating_grade / 2) if mark.rating_grade else None
            text = mark.comment_text
            source_url = _get_source_url(game)
            url = game.absolute_url
            line = [
                title,
                summary,
                world_rating,
                source_url,
                timestamp,
                my_rating,
                tags,
                text,
                url,
                "",
            ]
            ws.append(line)

    for status, label in [
        (ShelfType.COMPLETE, "听过的播客"),
        (ShelfType.PROGRESS, "在听的播客"),
        (ShelfType.WISHLIST, "想听的播客"),
    ]:
        ws = wb.create_sheet(title=label)
        shelf = user.shelf_manager.get_shelf(status)
        q = q_item_in_category(ItemCategory.Podcast)
        marks = shelf.members.all().filter(q).order_by("created_time")
        ws.append(heading)
        for mm in marks:
            mark = mm.mark
            podcast = mark.item
            title = podcast.title
            summary = ",".join(podcast.hosts or [])
            tags = ",".join(mark.tags)
            world_rating = (podcast.rating / 2) if podcast.rating else None
            timestamp = mark.created_time.strftime("%Y-%m-%d %H:%M:%S")
            my_rating = (mark.rating_grade / 2) if mark.rating_grade else None
            text = mark.comment_text
            source_url = _get_source_url(podcast)
            url = podcast.absolute_url
            line = [
                title,
                summary,
                world_rating,
                source_url,
                timestamp,
                my_rating,
                tags,
                text,
                url,
                "",
            ]
            ws.append(line)

    review_heading = [
        "标题",
        "评论对象",
        "链接",
        "创建时间",
        "我的评分",
        "类型",
        "内容",
        "评论对象原始链接",
        "评论对象NeoDB链接",
    ]
    for category, label in [
        (ItemCategory.Movie, "影评"),
        (ItemCategory.Book, "书评"),
        (ItemCategory.Music, "乐评"),
        (ItemCategory.Game, "游戏评论"),
        (ItemCategory.Podcast, "播客评论"),
    ]:
        ws = wb.create_sheet(title=label)
        q = q_item_in_category(category)
        reviews = (
            Review.objects.filter(owner=user.identity)
            .filter(q)
            .order_by("created_time")
        )
        ws.append(review_heading)
        for review in reviews:
            title = review.title
            target = "《" + review.item.title + "》"
            url = review.absolute_url
            timestamp = review.created_time.strftime("%Y-%m-%d %H:%M:%S")
            my_rating = None  # (mark.rating_grade / 2) if mark.rating_grade else None
            content = review.body
            target_source_url = _get_source_url(review.item)
            target_url = review.item.absolute_url
            line = [
                title,
                target,
                url,
                timestamp,
                my_rating,
                label,
                content,
                target_source_url,
                target_url,
            ]
            ws.append(line)

    wb.save(filename=filename)
    user.preference.export_status["marks_pending"] = False
    user.preference.export_status["marks_file"] = filename
    user.preference.export_status["marks_date"] = datetime.now().strftime(
        "%Y-%m-%d %H:%M"
    )
    user.preference.save(update_fields=["export_status"])
