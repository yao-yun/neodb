import os
import re
from datetime import datetime

import openpyxl
import pytz
from django.conf import settings
from loguru import logger
from markdownify import markdownify as md

from catalog.common import *
from catalog.common.downloaders import *
from catalog.models import *
from catalog.sites import DoubanBook, DoubanDrama, DoubanGame, DoubanMovie, DoubanMusic
from catalog.sites.douban import DoubanDownloader
from common.utils import GenerateDateUUIDMediaFilePath
from journal.models import *
from users.models import Task

_tz_sh = pytz.timezone("Asia/Shanghai")


def _fetch_remote_image(url):
    try:
        logger.info(f"fetching remote image {url}")
        imgdl = ProxiedImageDownloader(url)
        raw_img = imgdl.download().content
        ext = imgdl.extention
        f = GenerateDateUUIDMediaFilePath(f"x.{ext}", settings.MARKDOWNX_MEDIA_PATH)
        file = settings.MEDIA_ROOT + "/" + f
        local_url = settings.MEDIA_URL + f
        os.makedirs(os.path.dirname(file), exist_ok=True)
        with open(file, "wb") as binary_file:
            binary_file.write(raw_img)
        # logger.info(f'remote image saved as {local_url}')
        return local_url
    except Exception as e:
        logger.error(f"unable to fetch image", extra={"url": url, "exception": e})
        return url


class DoubanImporter(Task):
    class Meta:
        app_label = "journal"  # workaround bug in TypedModel

    TaskQueue = "import"
    DefaultMetadata = {
        "total": 0,
        "processed": 0,
        "skipped": 0,
        "imported": 0,
        "failed": 0,
        "mode": 0,
        "visibility": 0,
        "failed_urls": [],
        "file": None,
    }

    mark_sheet_config = {
        "想读": [ShelfType.WISHLIST],
        "在读": [ShelfType.PROGRESS],
        "读过": [ShelfType.COMPLETE],
        "想看": [ShelfType.WISHLIST],
        "在看": [ShelfType.PROGRESS],
        "看过": [ShelfType.COMPLETE],
        "想听": [ShelfType.WISHLIST],
        "在听": [ShelfType.PROGRESS],
        "听过": [ShelfType.COMPLETE],
        "想玩": [ShelfType.WISHLIST],
        "在玩": [ShelfType.PROGRESS],
        "玩过": [ShelfType.COMPLETE],
        "想看的舞台剧": [ShelfType.WISHLIST],
        "看过的舞台剧": [ShelfType.COMPLETE],
    }
    review_sheet_config = {
        "书评": [Edition],
        "影评": [Movie],
        "乐评": [Album],
        "剧评": [Performance],
        "游戏评论&攻略": [Game],
    }

    @classmethod
    def validate_file(cls, uploaded_file):
        try:
            wb = openpyxl.open(
                uploaded_file, read_only=True, data_only=True, keep_links=False
            )
            sheets = cls.mark_sheet_config.keys() | cls.review_sheet_config.keys()
            for name in sheets:
                if name in wb:
                    return True
        except Exception as e:
            logger.error(
                f"unable to validate excel file {uploaded_file}", extra={"exception": e}
            )
        return False

    mark_data = {}
    review_data = {}
    entity_lookup = {}

    def load_sheets(self):
        """Load data into mark_data / review_data / entity_lookup"""
        f = open(self.metadata["file"], "rb")
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True, keep_links=False)
        for data, config in [
            (self.mark_data, self.mark_sheet_config),
            (self.review_data, self.review_sheet_config),
        ]:
            for name in config:
                data[name] = []
                if name in wb:
                    logger.info(f"{self.user} parsing {name}")
                    for row in wb[name].iter_rows(min_row=2, values_only=True):
                        cells = [cell for cell in row]
                        if len(cells) > 6 and cells[0]:
                            data[name].append(cells)
        for sheet in self.mark_data.values():
            for cells in sheet:
                # entity_lookup["title|rating"] = [(url, time), ...]
                k = f"{cells[0]}|{cells[5]}"
                v = (cells[3], cells[4])
                if k in self.entity_lookup:
                    self.entity_lookup[k].append(v)
                else:
                    self.entity_lookup[k] = [v]
        self.metadata["total"] = sum(map(lambda a: len(a), self.mark_data.values()))
        self.metadata["total"] += sum(map(lambda a: len(a), self.review_data.values()))
        self.save()

    def guess_entity_url(self, title, rating, timestamp):
        k = f"{title}|{rating}"
        if k not in self.entity_lookup:
            return None
        v = self.entity_lookup[k]
        if len(v) > 1:
            v.sort(
                key=lambda c: abs(
                    timestamp
                    - (
                        datetime.strptime(c[1], "%Y-%m-%d %H:%M:%S")
                        if type(c[1]) == str
                        else c[1]
                    ).replace(tzinfo=_tz_sh)
                )
            )
        return v[0][0]
        # for sheet in self.mark_data.values():
        #     for cells in sheet:
        #         if cells[0] == title and cells[5] == rating:
        #             return cells[3]

    def run(self):
        logger.info(f"{self.user} import start")
        self.load_sheets()
        logger.info(f"{self.user} sheet loaded, {self.metadata['total']} lines total")
        for name, param in self.mark_sheet_config.items():
            self.import_mark_sheet(self.mark_data[name], param[0], name)
        for name, param in self.review_sheet_config.items():
            self.import_review_sheet(self.review_data[name], name)
        self.message = f"豆瓣标记和评论导入完成，共处理{self.metadata['total']}篇，已存在{self.metadata['skipped']}篇，新增{self.metadata['imported']}篇。"
        if len(self.metadata["failed_urls"]) > 0:
            self.message += f'导入时未能处理{len(self.metadata["failed_urls"])}个网址。'
        self.save()

    def import_mark_sheet(self, worksheet, shelf_type, sheet_name):
        prefix = f"{self.user} {sheet_name}|"
        if worksheet is None:  # or worksheet.max_row < 2:
            logger.warning(f"{prefix} empty sheet")
            return
        for cells in worksheet:
            if len(cells) < 6:
                continue
            # title = cells[0] or ""
            url = cells[3]
            time = cells[4]
            rating = cells[5]
            try:
                rating_grade = int(rating) * 2 if rating else None
            except Exception:
                rating_grade = None
            tags = cells[6] if len(cells) >= 7 else ""
            try:
                tags = tags.split(",") if tags else []
            except Exception:
                tags = []
            comment = cells[7] if len(cells) >= 8 else None
            self.metadata["processed"] += 1
            try:
                if type(time) == str:
                    time = datetime.strptime(time, "%Y-%m-%d %H:%M:%S")
                time = time.replace(tzinfo=_tz_sh)
            except Exception:
                time = None
            r = self.import_mark(url, shelf_type, comment, rating_grade, tags, time)
            if r == 1:
                self.metadata["imported"] += 1
            elif r == 2:
                self.metadata["skipped"] += 1
            self.save()

    def import_mark(self, url, shelf_type, comment, rating_grade, tags, time):
        """
        Import one mark: return 1: done / 2: skipped / None: failed
        """
        item = self.get_item_by_url(url)
        if not item:
            logger.warning(f"{self.user} | match/fetch {url} failed")
            return
        mark = Mark(self.user.identity, item)
        if self.metadata["mode"] == 0 and (
            mark.shelf_type == shelf_type
            or mark.shelf_type == ShelfType.COMPLETE
            or (
                mark.shelf_type in [ShelfType.PROGRESS, ShelfType.DROPPED]
                and shelf_type == ShelfType.WISHLIST
            )
        ):
            print("-", end="", flush=True)
            return 2
        mark.update(
            shelf_type,
            comment,
            rating_grade,
            tags,
            self.metadata["visibility"],
            created_time=time,
        )
        print("+", end="", flush=True)
        return 1

    def import_review_sheet(self, worksheet, sheet_name):
        prefix = f"{self.user} {sheet_name}|"
        if worksheet is None:  # or worksheet.max_row < 2:
            logger.warning(f"{prefix} empty sheet")
            return
        for cells in worksheet:
            if len(cells) < 6:
                continue
            title = cells[0]
            entity_title = (
                re.sub("^《", "", re.sub("》$", "", cells[1])) if cells[1] else ""
            )
            review_url = cells[2]
            time = cells[3]
            rating = cells[4]
            content = cells[6]
            self.metadata["processed"] += 1
            if time:
                if type(time) == str:
                    time = datetime.strptime(time, "%Y-%m-%d %H:%M:%S")
                time = time.replace(tzinfo=_tz_sh)
            else:
                time = None
            if not content:
                content = ""
            if not title:
                title = ""
            r = self.import_review(
                entity_title, rating, title, review_url, content, time
            )
            if r == 1:
                self.metadata["imported"] += 1
            elif r == 2:
                self.metadata["skipped"] += 1
            else:
                self.metadata["failed_urls"].append(review_url)
            self.save()

    def get_item_by_url(self, url):
        item = None
        if not url:
            logger.warning(f"URL empty")
            return None
        try:
            site = SiteManager.get_site_by_url(url)
            if not site:
                raise ValueError(f"URL unrecognized {url}")
            item = site.get_item()
            if not item:
                logger.info(f"fetching {url}")
                site.get_resource_ready()
                item = site.get_item()
            else:
                # logger.info(f"matched {url}")
                print(".", end="", flush=True)
        except DownloadError as e:
            if e.response_type == RESPONSE_CENSORSHIP:
                # avoid flood error log since there are too many
                logger.warning(f"fetching error: {url}", extra={"exception": e})
            else:
                logger.error(f"fetching error: {url}", extra={"exception": e})
        except Exception as e:
            logger.error(f"fetching error: {url}", extra={"exception": e})
        if item is None:
            self.metadata["failed_urls"].append(str(url))
        return item

    def is_douban_item_url(self, url):
        for cls in [
            DoubanBook,
            DoubanDrama,
            DoubanMovie,
            DoubanMusic,
            DoubanGame,
        ]:
            if cls.url_to_id(url):
                return True

    def import_review(self, entity_title, rating, title, review_url, content, time):
        """
        Import one review: return 1: done / 2: skipped / None: failed
        """
        prefix = f"{self.user} |"
        url = self.guess_entity_url(entity_title, rating, time)
        if url is None:
            logger.info(f"{prefix} fetching review {review_url}")
            try:
                h = DoubanDownloader(review_url).download().html()
                urls = h.xpath("//header[@class='main-hd']/a/@href")
                for u in urls:  # type:ignore
                    if self.is_douban_item_url(u):
                        url = u
                if not url:
                    logger.warning(
                        f"{prefix} fetching error {review_url} unable to locate entity url"
                    )
                    return
            except Exception:
                logger.error(f"{prefix} fetching review exception {review_url}")
                return
        item = self.get_item_by_url(url)
        if not item:
            logger.warning(f"{prefix} match/fetch {url} failed")
            return
        if (
            self.metadata["mode"] == 1
            and Review.objects.filter(owner=self.user.identity, item=item).exists()
        ):
            return 2
        content = re.sub(
            r'<span style="font-weight: bold;">([^<]+)</span>', r"<b>\1</b>", content
        )
        content = re.sub(r"(<img [^>]+>)", r"\1<br>", content)
        content = re.sub(
            r'<div class="image-caption">([^<]+)</div>', r"<br><i>\1</i><br>", content
        )
        content = md(content)
        content = re.sub(
            r"(?<=!\[\]\()([^)]+)(?=\))", lambda x: _fetch_remote_image(x[1]), content
        )
        params = {
            "created_time": time,
            "edited_time": time,
            "title": title,
            "body": content,
            "visibility": self.metadata["visibility"],
        }
        try:
            Review.objects.update_or_create(
                owner=self.user.identity, item=item, defaults=params
            )
        except Exception:
            logger.warning(f"{prefix} update multiple review {review_url}")
            r = (
                Review.objects.filter(owner=self.user.identity, item=item)
                .order_by("-created_time")
                .first()
            )
            if r:
                Review.objects.filter(pk=r.pk).update(**params)
        return 1
